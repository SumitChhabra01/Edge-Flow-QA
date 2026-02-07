"""Excel data reader for codeless tests."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class StepRecord:
    """Represents a single codeless step from Excel."""

    step: str
    action: str
    locator: Optional[str]
    value: Optional[str]
    expected: Optional[str]


@dataclass
class TestCaseRecord:
    """Represents a codeless test case entry."""

    test_case_id: str
    description: Optional[str]
    execute: bool


def load_steps_from_excel(path: str, sheet_name: Optional[str] = None) -> List[StepRecord]:
    """Load steps from an Excel sheet."""
    workbook = load_workbook(path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    steps: List[StepRecord] = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = _header_index(headers)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        steps.append(
            StepRecord(
                step=str(_get_cell(row, header_index.get("step", 0)) or "").strip(),
                action=str(_get_cell(row, header_index.get("action", 1)) or "").strip(),
                locator=_to_optional(_get_cell(row, header_index.get("locator", 2))),
                value=_to_optional(_get_cell(row, header_index.get("value", 3))),
                expected=_to_optional(_get_cell(row, header_index.get("expected", 4))),
            )
        )
    return steps


def load_flows_from_excel(path: str) -> Dict[str, List[StepRecord]]:
    """Load reusable flows from the FLOWS sheet."""
    workbook = load_workbook(path)
    if "FLOWS" not in workbook.sheetnames:
        return {}
    sheet = workbook["FLOWS"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = _header_index(headers)

    flows: Dict[str, List[StepRecord]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        flow_name = str(_get_cell(row, header_index.get("flowname", 0)) or "").strip()
        if not flow_name:
            continue
        record = StepRecord(
            step=str(_get_cell(row, header_index.get("step", 1)) or "").strip(),
            action=str(_get_cell(row, header_index.get("action", 2)) or "").strip(),
            locator=_to_optional(_get_cell(row, header_index.get("locator", 3))),
            value=_to_optional(_get_cell(row, header_index.get("value", 4))),
            expected=_to_optional(_get_cell(row, header_index.get("expected", 5))),
        )
        flows.setdefault(flow_name, []).append(record)

    for flow_name, records in flows.items():
        flows[flow_name] = _sort_steps(records)
    return flows


def load_testcases_from_excel(path: str) -> List[TestCaseRecord]:
    """Load test case list from the TestCases sheet."""
    workbook = load_workbook(path)
    if "TestCases" not in workbook.sheetnames:
        return []
    sheet = workbook["TestCases"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = _header_index(headers)

    testcases: List[TestCaseRecord] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        test_case_id = str(_get_cell(row, header_index.get("testcaseid", 0)) or "").strip()
        description = _to_optional(_get_cell(row, header_index.get("description", 1)))
        execute_raw = str(_get_cell(row, header_index.get("execute", 2)) or "").strip().lower()
        execute = execute_raw in {"yes", "y", "true", "1"}
        if test_case_id:
            testcases.append(TestCaseRecord(test_case_id=test_case_id, description=description, execute=execute))
    return testcases


def _to_optional(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _header_index(headers: List[object]) -> Dict[str, int]:
    normalized = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        key = str(header).strip().lower()
        normalized[key] = idx
    mapping = {
        "step": normalized.get("step"),
        "action": normalized.get("action"),
        "locator": normalized.get("locator", normalized.get("endpoint")),
        "value": normalized.get("value", normalized.get("payload")),
        "expected": normalized.get("expected"),
        "flowname": normalized.get("flowname"),
        "testcaseid": normalized.get("testcaseid"),
        "description": normalized.get("description"),
        "execute": normalized.get("execute"),
    }
    return {key: value for key, value in mapping.items() if value is not None}


def _get_cell(row, index: Optional[int]):
    if index is None:
        return None
    if index >= len(row):
        return None
    return row[index]


def _sort_steps(records: List[StepRecord]) -> List[StepRecord]:
    try:
        return sorted(records, key=lambda item: int(item.step))
    except (ValueError, TypeError):
        return records
