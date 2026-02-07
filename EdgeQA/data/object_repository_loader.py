"""Object repository loader for POM-style locators."""

from __future__ import annotations

import os
from typing import Dict, Tuple

from openpyxl import load_workbook

LocatorRepo = Dict[str, Dict[str, Tuple[str, str]]]
_CACHE: Dict[str, LocatorRepo] = {}


def load_object_repository(path: str) -> LocatorRepo:
    """Load ObjectRepository.xlsx and cache the results."""
    if not os.path.exists(path):
        return {}
    if path in _CACHE:
        return _CACHE[path]

    workbook = load_workbook(path)
    repository: LocatorRepo = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}
        if not {"locatorname", "locatortype", "locatorvalue"}.issubset(set(header_index.keys())):
            continue

        page_locators: Dict[str, Tuple[str, str]] = {}
        for row in rows[1:]:
            if not row or all(cell is None for cell in row):
                continue
            name = _to_text(_get_cell(row, header_index.get("locatorname")))
            locator_type = _to_text(_get_cell(row, header_index.get("locatortype")))
            locator_value = _to_text(_get_cell(row, header_index.get("locatorvalue")))
            if name and locator_type and locator_value:
                page_locators[name] = (locator_type, locator_value)
        if page_locators:
            repository[sheet_name] = page_locators

    _CACHE[path] = repository
    return repository


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _get_cell(row, index):
    if index is None:
        return None
    if index >= len(row):
        return None
    return row[index]
