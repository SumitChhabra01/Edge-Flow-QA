"""Excel-driven DSL executor."""

from __future__ import annotations

import os
from dataclasses import dataclass
from typing import Dict, List, Optional

from openpyxl import load_workbook

from core.api_client import ApiClient
from core.commands.api_call import ApiCallCommand
from core.commands.call_flow import CallFlowCommand
from core.commands.click import ClickCommand
from core.commands.open_url import OpenUrlCommand
from core.commands.store_response import StoreResponseCommand
from core.commands.take_screenshot import TakeScreenshotCommand
from core.commands.type_text import TypeTextCommand
from core.commands.verify_status import VerifyStatusCommand
from core.commands.verify_text import VerifyTextCommand
from core.commands.verify_visible import VerifyVisibleCommand
from core.conditions.condition_evaluator import ConditionEvaluator
from core.engine.command_registry import CommandRegistry
from core.engine.context_store import ContextStore
from core.engine.hook_executor import HookExecutor
from core.locator.locator_loader import load_locator_repository
from core.locator.locator_resolver import LocatorResolver
from core.playwright_manager import PlaywrightManager
from utils.config_loader import load_config
from utils.file_utils import ensure_dirs, resolve_path
from utils.logger import get_logger
from utils.dsl_report import TestRecord, clean_screenshots, safe_name, summarize_error, write_report


@dataclass
class StepRow:
    """Represents a DSL step row."""

    seq: str
    execute: str
    command: str
    target: str
    data: str
    condition: str
    store: str
    failure_category: str


class DslExecutor:
    """Execute Excel DSL tests."""

    def __init__(
        self,
        root_dir: str,
        locator_repo_path: Optional[str] = None,
        flows_dir: Optional[str] = None,
    ) -> None:
        self.root_dir = root_dir
        self.config = load_config(root_dir)
        self.logger = get_logger("edgeqa_dsl", logs_dir=resolve_path(root_dir, "logs"))
        self.locator_repo_path = locator_repo_path or resolve_path(root_dir, "InputSheet", "LocatorRepository.xlsx")
        self.flows_dir = flows_dir or resolve_path(root_dir, "flows")
        self._flow_depth = 0
        self._max_flow_depth = 3

        ensure_dirs([self.flows_dir])

        self.registry = CommandRegistry()
        self._register_commands()

    def execute(self, testcases_path: str) -> None:
        """Execute test cases from the provided Excel file."""
        workbook = load_workbook(testcases_path)
        testcases_sheet = workbook["TestCases"]
        testcases = self._read_testcases(testcases_sheet)

        reports_dir = resolve_path(self.root_dir, "reports")
        artifacts_dir = resolve_path(reports_dir, "artifacts")
        ensure_dirs([reports_dir, artifacts_dir])
        clean_screenshots(self.root_dir)

        browser_name = self.config["browser_name"]
        browser_config = self.config["browser"]
        env = self.config["environment"]

        manager = PlaywrightManager(browser_name, browser_config, artifacts_dir, record_video=bool(self.config["config"]["video"]["enabled"]))
        page = manager.start()

        api_client = ApiClient(
            base_url=env.get("api_base_url", ""),
            timeout_ms=self.config["config"]["timeouts"]["api"],
            playwright=manager.playwright,
        )
        api_client.start()

        try:
            context = self._build_context(page, api_client)
            context.set("screenshot_func", manager.screenshot_on_failure)
            hook_executor = HookExecutor(self)
            report_records: List[TestRecord] = []
            errors: List[Exception] = []
            for tc in testcases:
                if tc["execute"] != "Y":
                    continue
                self.logger.info("DSL TestCase: %s", tc["id"])
                steps_log: List[str] = []
                failed_step = None
                root_cause = None
                screenshot_path = None
                status = "PASSED"
                failure_entries: List[dict] = []
                try:
                    hook_executor.execute_hook(tc.get("before_hook"), workbook, context=context)
                    self.execute_steps_sheet(
                        workbook,
                        tc["steps_sheet"],
                        context=context,
                        steps_log=steps_log,
                        failure_entries=failure_entries,
                    )
                    hook_executor.execute_hook(tc.get("after_hook"), workbook, context=context)
                except Exception as exc:  # noqa: BLE001
                    status = "FAILED"
                    root_cause = str(exc)
                    failed_step = steps_log[-1] if steps_log else None
                    filename = f"{safe_name(tc['id'])}"
                    screenshot_path = manager.screenshot_on_failure(filename)
                    errors.append(exc)
                finally:
                    if failure_entries and status != "FAILED":
                        status = "FAILED"
                    report_records.append(
                        TestRecord(
                            nodeid=tc["id"],
                            test_name=tc["id"],
                            status=status,
                            root_cause=root_cause,
                            failed_step=failed_step,
                            steps=steps_log,
                            stderr="",
                            error_snippet=summarize_error(root_cause or ""),
                            screenshot_path=screenshot_path,
                            failure_entries=failure_entries,
                        )
                    )
            write_report(self.root_dir, report_records)
            if errors:
                raise errors[0]
        finally:
            api_client.stop()
            manager.stop()

    def execute_steps_sheet(
        self,
        workbook,
        sheet_name: str,
        context: Optional[ContextStore] = None,
        steps_log: Optional[List[str]] = None,
        failure_entries: Optional[List[dict]] = None,
    ) -> None:
        """Execute a sheet containing DSL steps."""
        sheet = workbook[sheet_name]
        steps = self._read_steps(sheet)
        ctx = context or self._build_context(None, None)

        for step in steps:
            self._execute_step(step, ctx, steps_log=steps_log, failure_entries=failure_entries)

    def _execute_step(
        self,
        step: StepRow,
        context: ContextStore,
        steps_log: Optional[List[str]] = None,
        failure_entries: Optional[List[dict]] = None,
    ) -> None:
        if step.execute and step.execute.upper() not in {"Y", "YES", "TRUE", "1"}:
            return
        command_name = step.command.upper()
        command = self.registry.get(command_name)

        target = context.resolve(step.target)
        data = context.resolve(step.data)

        if target and "." in target and command_name not in {"OPEN_URL", "API_CALL"}:
            resolver: LocatorResolver = context.get("locator_resolver")
            target_resolved = resolver.resolve(target)
        else:
            target_resolved = target

        condition_eval = context.get("condition_evaluator")
        condition_result = condition_eval.evaluate(step.condition, target_resolved, context)
        if not condition_result.should_execute:
            self.logger.info(
                "DSL step skipped | command=%s target=%s resolved=%s condition=%s",
                command_name,
                target,
                target_resolved,
                step.condition,
            )
            return

        attempts = max(1, condition_result.retry_count + 1)
        last_error = None
        for attempt in range(attempts):
            try:
                self.logger.info(
                    "DSL step | command=%s target=%s resolved=%s data=%s condition=%s store=%s",
                    command_name,
                    target,
                    target_resolved,
                    data,
                    step.condition,
                    step.store,
                )
                if steps_log is not None:
                    steps_log.append(f"{command_name} | {target or ''} | {data or ''}")
                result = command.execute(target_resolved, data, context)
                if step.store:
                    context.set(step.store, result)
                return
            except Exception as exc:  # noqa: BLE001
                last_error = exc
                self.logger.error("DSL step failed (attempt %s/%s): %s", attempt + 1, attempts, str(exc))
        if last_error:
            failure_category = (step.failure_category or "STOP_ON_FAILURE").upper()
            screenshot_func = context.get("screenshot_func")
            screenshot_path = None
            if screenshot_func:
                screenshot_path = screenshot_func(f"{safe_name(step.command)}_{safe_name(step.seq or 'step')}")
            if failure_entries is not None:
                failure_entries.append(
                    {
                        "step": f"{command_name} | {target or ''} | {data or ''}",
                        "category": failure_category,
                        "error": str(last_error),
                        "screenshot": screenshot_path,
                    }
                )
            if failure_category == "CONTINUE_ON_FAILURE":
                return
        if last_error:
            raise last_error

    def _register_commands(self) -> None:
        self.registry.register(OpenUrlCommand.name, OpenUrlCommand())
        self.registry.register(ClickCommand.name, ClickCommand())
        self.registry.register(TypeTextCommand.name, TypeTextCommand())
        self.registry.register(VerifyTextCommand.name, VerifyTextCommand())
        self.registry.register(VerifyVisibleCommand.name, VerifyVisibleCommand())
        self.registry.register(CallFlowCommand.name, CallFlowCommand())
        self.registry.register(ApiCallCommand.name, ApiCallCommand())
        self.registry.register(VerifyStatusCommand.name, VerifyStatusCommand())
        self.registry.register(StoreResponseCommand.name, StoreResponseCommand())
        self.registry.register(TakeScreenshotCommand.name, TakeScreenshotCommand())

    def _build_context(self, page, api_client) -> ContextStore:
        env = self.config["environment"]
        context = ContextStore(
            {
                "BASE_URL": env.get("base_url", ""),
                "TIMEOUT_MS": self.config["config"]["timeouts"]["default"],
                "ARTIFACTS_DIR": resolve_path(self.root_dir, "reports", "artifacts"),
            }
        )
        if page is not None:
            context.set("page", page)
        if api_client is not None:
            context.set("api_client", api_client)

        repository = load_locator_repository(self.locator_repo_path)
        context.set("locator_resolver", LocatorResolver(repository))
        context.set("condition_evaluator", ConditionEvaluator(timeout_ms=self.config["config"]["timeouts"]["default"]))
        context.set("flow_executor", lambda name, params: self._execute_flow(name, params, context))
        return context

    def _execute_flow(self, flow_name: str, params: Dict[str, str], context: ContextStore) -> None:
        if not flow_name:
            raise ValueError("CALL_FLOW requires a flow name.")
        if self._flow_depth >= self._max_flow_depth:
            raise ValueError("Flow depth exceeded.")
        flow_path = resolve_path(self.flows_dir, f"{flow_name}.flow.xlsx")
        if not os.path.exists(flow_path):
            raise ValueError(f"Missing flow: {flow_name}")
        self._flow_depth += 1
        try:
            workbook = load_workbook(flow_path)
            sheet_name = workbook.sheetnames[0]
            old_values = {}
            for key, value in params.items():
                old_values[key] = context.get(key)
                context.set(key, value)
            self.execute_steps_sheet(workbook, sheet_name, context=context)
            for key, value in old_values.items():
                context.set(key, value)
        finally:
            self._flow_depth -= 1

    def _read_testcases(self, sheet) -> List[Dict[str, str]]:
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {name: idx for idx, name in enumerate(headers)}
        execute_idx = _pick_header(header_index, ["Execute (Y/N)", "Execute"])
        before_idx = _pick_header(header_index, ["BeforeHook"])
        after_idx = _pick_header(header_index, ["AfterHook"])
        steps_idx = _pick_header(header_index, ["StepsSheet"])
        tc_id_idx = _pick_header(header_index, ["TestCaseID"])
        cases: List[Dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            tc_id = _cell(row, tc_id_idx)
            execute = _cell(row, execute_idx).upper() or "N"
            steps_sheet = _cell(row, steps_idx)
            cases.append(
                {
                    "id": tc_id,
                    "execute": execute,
                    "before_hook": _cell(row, before_idx),
                    "after_hook": _cell(row, after_idx),
                    "steps_sheet": steps_sheet,
                }
            )
        return cases

    def _read_steps(self, sheet) -> List[StepRow]:
        headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {name: idx for idx, name in enumerate(headers)}
        execute_idx = _pick_header(header_index, ["Execute (Y/N)", "Execute"])
        failure_idx = _pick_header(header_index, ["Failure Category", "FailureCategory"])
        steps = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            steps.append(
                StepRow(
                    seq=_cell(row, header_index.get("Seq")),
                    execute=_cell(row, execute_idx),
                    command=_cell(row, header_index.get("COMMAND")),
                    target=_cell(row, header_index.get("TARGET")),
                    data=_cell(row, header_index.get("DATA")),
                    condition=_cell(row, header_index.get("CONDITION")),
                    store=_cell(row, header_index.get("STORE")),
                    failure_category=_cell(row, failure_idx),
                )
            )
        return steps


def _cell(row, index) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def _pick_header(header_index: Dict[str, int], names: List[str]) -> Optional[int]:
    for name in names:
        if name in header_index:
            return header_index[name]
    return None
