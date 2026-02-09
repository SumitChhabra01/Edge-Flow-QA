"""Locator repository loader."""

from __future__ import annotations

import os
from typing import Dict, Tuple

from openpyxl import load_workbook

LocatorRepo = Dict[str, Dict[str, Tuple[str, str, str, str]]]
_CACHE: Dict[str, LocatorRepo] = {}


def load_locator_repository(path: str) -> LocatorRepo:
    """Load locator repository Excel and cache by path."""
    if not os.path.exists(path):
        return {}
    if path in _CACHE:
        return _CACHE[path]

    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required = {"page", "name", "primary", "secondary", "type"}
    if not required.issubset(set(header_index.keys())):
        return {}

    repo: LocatorRepo = {}
    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        page = _cell(row, header_index.get("page"))
        name = _cell(row, header_index.get("name"))
        primary = _cell(row, header_index.get("primary"))
        secondary = _cell(row, header_index.get("secondary"))
        loc_type = _cell(row, header_index.get("type"))
        if page and name and primary and loc_type:
            repo.setdefault(page, {})[name] = (primary, secondary, loc_type, page)

    _CACHE[path] = repo
    return repo


def _cell(row, index) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()
